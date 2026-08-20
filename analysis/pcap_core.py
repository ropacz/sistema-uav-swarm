"""Decodificação e correspondência de pacotes PCAPNG do ECHOSAR-Net.

Este módulo é compartilhado pelo relatório batch e pelos testes. A interface
pública para gerar planilhas é ``pcap_batch_to_spreadsheet.py``.

Exemplo de mensagem original produzida pelo ECHOSAR-Net::

    PositionUpdateChunk(
        messageId="team0-pos-1",
        senderId="team0",
        ipAddress="10.0.0.2",
        sequenceNumber=1,
        positionX=540,
        positionY=500,
        positionZ=1.5
    )

No PCAP ela é observada desta forma::

    10.0.0.2:5000 -> 255.255.255.255:5000
    UDP payload = 160 bytes -> PositionUpdate

Outros exemplos reconhecidos::

    10.0.0.1:5000 -> 10.0.0.2:5000
    UDP payload = 320 bytes -> VictimAlert

    10.0.0.2:5000 -> 10.0.0.1:5000
    UDP payload = 96 bytes -> VictimAck

O payload começa com ``ECHO``, versão e código do tipo. Os campos internos,
como ``messageId``, ``alertId`` e sequência, são decodificados diretamente do
PCAP. O tamanho só é usado como compatibilidade para capturas antigas.
"""

from __future__ import annotations

import ipaddress
import struct
from collections import Counter
from pathlib import Path

from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


# Tamanhos dos payloads UDP configurados nas aplicações. Eles não incluem os
# 8 bytes do cabeçalho UDP.
MESSAGE_TYPES = {
    96: "VictimAck",
    160: "PositionUpdate",
    320: "VictimAlert",
}

WIRE_MESSAGE_TYPES = {
    1: "PositionUpdate",
    2: "VictimAlert",
    3: "VictimAck",
}


def decode_echosar_payload(payload: bytes) -> dict:
    """Decodifica o formato binário ECHOSAR iniciado pela assinatura ``ECHO``.

    Capturas antigas, ou payloads que não pertencem ao ECHOSAR, retornam um
    dicionário vazio e continuam sendo classificados pelo método legado.
    """

    if len(payload) < 8 or payload[:4] != b"ECHO":
        return {}
    version, message_code, body_length = struct.unpack_from("!BBH", payload, 4)
    if version != 1 or message_code not in WIRE_MESSAGE_TYPES:
        return {}
    if 8 + body_length > len(payload):
        return {}

    body = payload[8:8 + body_length]
    offset = 0

    def take(length: int) -> bytes:
        nonlocal offset
        if offset + length > len(body):
            raise ValueError("campo ultrapassa o corpo ECHOSAR")
        value = body[offset:offset + length]
        offset += length
        return value

    def text() -> str:
        length = struct.unpack("!H", take(2))[0]
        return take(length).decode("utf-8", errors="replace")

    def uint32() -> int:
        return struct.unpack("!I", take(4))[0]

    def int32() -> int:
        return struct.unpack("!i", take(4))[0]

    def int64() -> int:
        return struct.unpack("!q", take(8))[0]

    def number() -> float:
        return struct.unpack("!d", take(8))[0]

    decoded = {
        "wire_magic": "ECHO",
        "wire_version": version,
        "message_type": WIRE_MESSAGE_TYPES[message_code],
        "encoded_body_bytes": body_length,
    }
    try:
        if message_code == 1:
            decoded.update(
                message_id=text(), sender_id=text(), sender_type=text(),
                advertised_ip=text(), waypoint_id=int32(), position_x=number(),
                position_y=number(), position_z=number(), sequence_number=int64(),
                message_timestamp=number(), operational_state=text(),
            )
        elif message_code == 2:
            decoded.update(
                alert_id=text(), message_id=text(), victim_id=text(),
                origin_drone_id=text(), origin_drone_address=text(),
                victim_position_x=number(), victim_position_y=number(),
                victim_position_z=number(), drone_position_x=number(),
                drone_position_y=number(), drone_position_z=number(),
                waypoint_id=int32(), sequence_number=int64(),
                attempt_number=int32(), generation_timestamp=number(),
                transmission_timestamp=number(), time_to_live=number(),
            )
        else:
            decoded.update(
                alert_id=text(), received_message_id=text(), victim_id=text(),
                team_id=text(), origin_drone_id=text(),
                reception_timestamp=number(), ack_timestamp=number(),
            )
    except (UnicodeError, ValueError, struct.error):
        return {}
    return decoded


def parse_options(data: bytes, endian: str) -> dict[int, list[bytes]]:
    """Lê opções de um bloco PCAPNG, respeitando o alinhamento de 32 bits."""

    options: dict[int, list[bytes]] = {}
    offset = 0
    while offset + 4 <= len(data):
        code, length = struct.unpack_from(endian + "HH", data, offset)
        offset += 4
        if code == 0:
            break
        value = data[offset:offset + length]
        options.setdefault(code, []).append(value)
        offset += (length + 3) & ~3
    return options


def timestamp_scale(options: dict[int, list[bytes]]) -> float:
    """Converte a opção PCAPNG if_tsresol em uma escala de segundos."""

    # A resolução padrão é 10^-6 segundo quando a opção não está presente.
    value = options.get(9, [b"\x06"])[0][0]
    if value & 0x80:
        return 2.0 ** -(value & 0x7F)
    return 10.0 ** -value


def iter_pcapng_packets(path: Path):
    """Percorre os pacotes sem depender de Scapy, PyShark ou tshark.

    Para cada Enhanced Packet Block, retorna o tempo em segundos, o ID da
    interface e os bytes capturados do quadro IEEE 802.11.
    """

    data = path.read_bytes()
    offset = 0
    endian = "<"
    scales: list[float] = []
    interface_ips: list[str | None] = []
    interface_names: list[str] = []

    while offset + 12 <= len(data):
        block_type_bytes = data[offset:offset + 4]
        if block_type_bytes == b"\x0a\x0d\x0d\x0a":
            byte_order_magic = data[offset + 8:offset + 12]
            endian = "<" if byte_order_magic == b"\x4d\x3c\x2b\x1a" else ">"
            scales = []
            interface_ips = []
            interface_names = []

        block_type, block_length = struct.unpack_from(endian + "II", data, offset)
        if block_length < 12 or offset + block_length > len(data):
            raise ValueError(f"Bloco PCAPNG inválido em {path}: offset {offset}")
        body = data[offset + 8:offset + block_length - 4]

        # Interface Description Block: contém a resolução dos timestamps.
        if block_type == 1 and len(body) >= 8:
            options = parse_options(body[8:], endian)
            scales.append(timestamp_scale(options))
            ipv4_option = options.get(4, [b""])[0]
            interface_ips.append(
                str(ipaddress.ip_address(ipv4_option[:4]))
                if len(ipv4_option) >= 4 else None
            )
            name_option = options.get(2, [b""])[0]
            interface_names.append(name_option.decode("utf-8", errors="replace"))
        # Enhanced Packet Block: contém um quadro enviado ou recebido.
        elif block_type == 6 and len(body) >= 20:
            interface_id, ts_high, ts_low, captured_length, _ = struct.unpack_from(
                endian + "IIIII", body, 0
            )
            packet = body[20:20 + captured_length]
            scale = scales[interface_id] if interface_id < len(scales) else 1e-6
            timestamp = ((ts_high << 32) | ts_low) * scale
            options_offset = 20 + ((captured_length + 3) & ~3)
            options = parse_options(body[options_offset:], endian)
            flags_option = options.get(2, [b"\x00\x00\x00\x00"])[0]
            flags = struct.unpack_from(endian + "I", flags_option)[0]
            direction = {1: "inbound", 2: "outbound"}.get(flags & 0x3, "unknown")
            interface_ip = (
                interface_ips[interface_id]
                if interface_id < len(interface_ips) else None
            )
            interface_name = (
                interface_names[interface_id]
                if interface_id < len(interface_names) else ""
            )
            yield (
                timestamp,
                interface_id,
                packet,
                direction,
                interface_ip,
                interface_name,
            )

        offset += block_length


def find_ipv4_udp(packet: bytes) -> dict | None:
    """Localiza e decodifica IPv4/UDP dentro de um quadro IEEE 802.11.

    Como os cabeçalhos 802.11/LLC podem variar, procura-se um cabeçalho IPv4
    válido em vez de assumir uma posição fixa. ARP e ACK puro do Wi-Fi são
    ignorados porque não contêm IPv4/UDP.
    """

    for offset in range(0, max(0, len(packet) - 28)):
        first = packet[offset]
        if first >> 4 != 4:
            continue
        ihl = (first & 0x0F) * 4
        if ihl < 20 or offset + ihl + 8 > len(packet):
            continue
        total_length = struct.unpack_from("!H", packet, offset + 2)[0]
        if total_length < ihl + 8 or offset + total_length > len(packet):
            continue
        if packet[offset + 9] != 17:
            continue

        udp_offset = offset + ihl
        src_port, dst_port, udp_length = struct.unpack_from(
            "!HHH", packet, udp_offset
        )
        if udp_length < 8 or udp_offset + udp_length > len(packet):
            continue

        udp_payload = packet[udp_offset + 8:udp_offset + udp_length]
        return {
            "source_ip": str(ipaddress.ip_address(packet[offset + 12:offset + 16])),
            "destination_ip": str(
                ipaddress.ip_address(packet[offset + 16:offset + 20])
            ),
            "ip_id": struct.unpack_from("!H", packet, offset + 4)[0],
            "ttl": packet[offset + 8],
            "source_port": src_port,
            "destination_port": dst_port,
            "udp_payload_bytes": udp_length - 8,
            "frame_bytes": len(packet),
            "udp_payload": udp_payload,
        }
    return None


def load_capture(path: Path, endpoint: str) -> list[dict]:
    """Carrega eventos UDP e atribui um tipo de mensagem a cada pacote."""

    events = []
    for number, (
        timestamp,
        interface_id,
        packet,
        direction,
        interface_ip,
        interface_name,
    ) in enumerate(
        iter_pcapng_packets(path), start=1
    ):
        decoded = find_ipv4_udp(packet)
        if decoded is None:
            continue
        # AODV usa UDP/654. Para UDP/5000, o cabeçalho ECHO tem prioridade;
        # tamanhos são apenas o fallback para capturas antigas.
        wire_fields = decode_echosar_payload(decoded.pop("udp_payload"))
        if decoded["source_port"] == 654 or decoded["destination_port"] == 654:
            message_type = "AODV"
        elif wire_fields:
            message_type = wire_fields["message_type"]
        else:
            message_type = MESSAGE_TYPES.get(
                decoded["udp_payload_bytes"], "OtherUdp"
            )
        decoded.update(
            {
                "capture": endpoint,
                "packet_number": number,
                "time_seconds": timestamp,
                "interface_id": interface_id,
                "interface_ip": interface_ip,
                "interface_name": interface_name,
                "direction": direction,
                "message_type": message_type,
            }
        )
        decoded.update(wire_fields)
        events.append(decoded)
    return events


def packet_key(event: dict) -> tuple:
    """Cria a identidade usada para procurar o pacote na outra captura.

    Mensagens novas usam os identificadores funcionais gravados no payload,
    por exemplo ``messageId`` e ``sequenceNumber``. Isso é mais auditável que
    inferir identidade por tamanho ou pelo ID IPv4. Capturas antigas mantêm o
    fallback baseado nos cabeçalhos de rede::

        ("10.0.0.2", "255.255.255.255", 3, 5000, 5000, 160)
    """

    common = (
        "ECHO-v1",
        event.get("message_type"),
        event["source_ip"],
        event["destination_ip"],
    )
    if event.get("wire_magic") == "ECHO":
        if event.get("message_type") == "VictimAck":
            identity = event.get("received_message_id") or event.get("alert_id")
            if identity:
                return common + (identity, event.get("alert_id", ""))
        else:
            identity = event.get("message_id")
            if identity:
                return common + (
                    identity,
                    event.get("sequence_number", ""),
                    event.get("attempt_number", ""),
                )

    return (
        event["source_ip"],
        event["destination_ip"],
        event["ip_id"],
        event["source_port"],
        event["destination_port"],
        event["udp_payload_bytes"],
    )


def packet_identity_method(event: dict) -> str:
    """Informa na auditoria como a identidade do pacote foi determinada."""

    key = packet_key(event)
    return "echo_fields" if key and key[0] == "ECHO-v1" else "network_headers"


def is_group_destination(address: str) -> bool:
    """Indica broadcast limitado ou endereço multicast IPv4."""

    ip = ipaddress.ip_address(address)
    return address == "255.255.255.255" or ip.is_multicast


def compare_direction(
    source_events: list[dict], destination_events: list[dict], source_ip: str
) -> list[dict]:
    """Compara o que saiu de uma origem com o observado no outro nó.

    Um pacote é entregue quando sua chave existe nas duas capturas. O Counter
    trata corretamente retransmissões ou pacotes repetidos.
    """

    received = Counter(
        packet_key(event)
        for event in destination_events
        if event["source_ip"] == source_ip and event["direction"] != "outbound"
    )
    comparisons = []
    for event in source_events:
        if event["source_ip"] != source_ip or event["direction"] == "inbound":
            continue
        key = packet_key(event)
        delivered = received[key] > 0
        if delivered:
            received[key] -= 1
        comparisons.append(
            {
                "message_type": event["message_type"],
                "metric_scope": (
                    "broadcast_reception_opportunity"
                    if is_group_destination(event["destination_ip"])
                    else "unicast_packet_delivery"
                ),
                "expected_receiver_basis": "specified_capture_pair",
                "identity_method": packet_identity_method(event),
                "source_ip": event["source_ip"],
                "destination_ip": event["destination_ip"],
                "ip_id": event["ip_id"],
                "source_port": event["source_port"],
                "destination_port": event["destination_port"],
                "udp_payload_bytes": event["udp_payload_bytes"],
                "wire_magic": event.get("wire_magic", ""),
                "wire_version": event.get("wire_version", ""),
                "message_id": event.get("message_id", ""),
                "received_message_id": event.get("received_message_id", ""),
                "alert_id": event.get("alert_id", ""),
                "sequence_number": event.get("sequence_number", ""),
                "attempt_number": event.get("attempt_number", ""),
                "sent_time_seconds": event["time_seconds"],
                "received": "Sim" if delivered else "Não",
                "lost": 0 if delivered else 1,
            }
        )
    return comparisons


def format_workbook(path: Path) -> None:
    """Aplica filtros, cabeçalhos, larguras e percentuais à planilha."""

    from openpyxl import load_workbook

    workbook = load_workbook(path)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for cell in sheet[1]:
            cell.font = Font(color="FFFFFF", bold=True)
            cell.fill = header_fill
        for column in sheet.columns:
            width = min(max(len(str(cell.value or "")) for cell in column) + 2, 32)
            sheet.column_dimensions[get_column_letter(column[0].column)].width = width
    # Qualquer aba que tenha PDR/perda recebe apresentação percentual.
    for sheet in workbook.worksheets:
        headers = {cell.value: cell.column for cell in sheet[1]}
        percentage_names = {
            name for name in headers
            if isinstance(name, str)
            and (name == "pdr" or name == "loss_rate"
                 or name.startswith("pdr_") or name.startswith("loss_"))
        }
        for name in percentage_names:
            for row in range(2, sheet.max_row + 1):
                sheet.cell(row, headers[name]).number_format = "0.00%"
    workbook.save(path)
