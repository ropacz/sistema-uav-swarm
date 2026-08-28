#!/usr/bin/env python3
"""Planilha com fórmulas de verdade (COUNTIFS/SUMIFS/STDEV), não valores
pré-calculados pelo Python — pra quem abrir conferir, célula por célula, como
o atendimento/perda de cada seed foi obtido a partir da aba `Alertas`.

Três abas:
    Alertas         — cópia da aba de mesmo nome de `alert_sheet.py`
    PorSeed         — 1 linha por (config, numTeams, seed): gerados/confirmados/
                      sem_entrega via COUNTIFS/SUMIFS, atendimento/perda % por
                      divisão de célula
    ResumoFormulas  — 1 linha por (config, numTeams): média e desvio-padrão das
                      30 linhas de PorSeed daquela célula, via AVERAGE/STDEV

Não recalcula nada em Python além de decidir quais fórmulas escrever e em
qual célula — os números saem todos da avaliação das fórmulas pelo programa
que abrir o arquivo.
"""

from __future__ import annotations

from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter

from analysis.reports.alert_sheet import COLUMNS, load

REPOSITORY_ROOT = Path(__file__).resolve().parents[2]
OUTPUT = REPOSITORY_ROOT / "analysis/tables/atendimento_formulas.xlsx"


def write_alerts_sheet(workbook: openpyxl.Workbook, table) -> dict[str, str]:
    """Copia a aba Alertas e devolve o mapa nome-de-coluna -> letra."""
    sheet = workbook.active
    sheet.title = "Alertas"
    sheet.append(COLUMNS)
    for row in table[COLUMNS].itertuples(index=False):
        sheet.append(list(row))
    return {name: get_column_letter(index + 1) for index, name in enumerate(COLUMNS)}


def write_per_seed_sheet(workbook: openpyxl.Workbook, table,
                          alerts_column: dict[str, str]) -> list[tuple[str, int, int, int]]:
    """Uma linha por (config, numTeams, seed); devolve os blocos contíguos
    (config, numTeams, primeira_linha, última_linha) pra ResumoFormulas usar.
    """
    sheet = workbook.create_sheet("PorSeed")
    columns = ["config", "numTeams", "seed", "gerados", "confirmados",
               "sem_entrega", "atendimento_pct", "perda_pct"]
    sheet.append(columns)
    column_letter = {name: get_column_letter(index + 1)
                      for index, name in enumerate(columns)}

    def alerts_range(name: str) -> str:
        letter = alerts_column[name]
        return f"Alertas!${letter}:${letter}"

    blocks = []
    row = 2
    for config in sorted(table["config"].unique()):
        teams_values = sorted(table.loc[table["config"] == config, "numTeams"].unique())
        for teams in teams_values:
            first_row = row
            seeds = sorted(table.loc[
                (table["config"] == config) & (table["numTeams"] == teams), "seed"
            ].unique())
            for seed in seeds:
                generated = (
                    f"=COUNTIFS({alerts_range('config')},\"{config}\","
                    f"{alerts_range('numTeams')},{teams},"
                    f"{alerts_range('seed')},{seed})"
                )
                confirmed = (
                    f"=SUMIFS({alerts_range('acknowledged')},"
                    f"{alerts_range('config')},\"{config}\","
                    f"{alerts_range('numTeams')},{teams},"
                    f"{alerts_range('seed')},{seed})"
                )
                undelivered = (
                    f"=COUNTIFS({alerts_range('config')},\"{config}\","
                    f"{alerts_range('numTeams')},{teams},"
                    f"{alerts_range('seed')},{seed},"
                    f"{alerts_range('delivered')},0)"
                )
                generated_cell = f"{column_letter['gerados']}{row}"
                confirmed_cell = f"{column_letter['confirmados']}{row}"
                undelivered_cell = f"{column_letter['sem_entrega']}{row}"
                sheet.append([
                    config, teams, seed,
                    generated, confirmed, undelivered,
                    f"={confirmed_cell}/{generated_cell}*100",
                    f"={undelivered_cell}/{generated_cell}*100",
                ])
                row += 1
            blocks.append((config, teams, first_row, row - 1))
    return blocks


def write_summary_sheet(workbook: openpyxl.Workbook,
                         blocks: list[tuple[str, int, int, int]]) -> None:
    sheet = workbook.create_sheet("ResumoFormulas")
    sheet.append(["config", "numTeams", "pares",
                  "atendimento_media_pct", "atendimento_desvio_pct",
                  "perda_media_pct", "perda_desvio_pct"])
    for config, teams, first_row, last_row in blocks:
        attendance_range = f"PorSeed!G{first_row}:G{last_row}"
        loss_range = f"PorSeed!H{first_row}:H{last_row}"
        sheet.append([
            config, teams, last_row - first_row + 1,
            f"=AVERAGE({attendance_range})", f"=STDEV({attendance_range})",
            f"=AVERAGE({loss_range})", f"=STDEV({loss_range})",
        ])


def main() -> None:
    table = load()
    workbook = openpyxl.Workbook()
    alerts_column = write_alerts_sheet(workbook, table)
    blocks = write_per_seed_sheet(workbook, table, alerts_column)
    write_summary_sheet(workbook, blocks)

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(OUTPUT)
    print(f"gerado: {OUTPUT.relative_to(REPOSITORY_ROOT)}")
    print(f"  Alertas: {len(table)} linhas")
    print(f"  PorSeed: {blocks[-1][3] - 1} linhas, {len(blocks)} células (config x numTeams)")
    print(f"  ResumoFormulas: {len(blocks)} linhas")


if __name__ == "__main__":
    main()
